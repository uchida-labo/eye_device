from email.mime import base
import cv2
import openpyxl
import pygame
import time
import matplotlib.pyplot as plt
import schedule

pygame.mixer.init()
pygame.mixer.music.load('.mp3')

def judge_square(trim_frame, xmin, xmax, ymin, ymax):
    video_size = (xmax - xmin)*(ymax - ymin)
    white = cv2.countNonZero(trim_frame)
    black = video_size - white
    white_ratio = (white/video_size) * 100
    black_ratio = (black/video_size) * 100

    return white_ratio, black_ratio

def filter(frame, xmin, xmax, ymin, ymax):

    # GaussianFilter processing (smoothing)
    fil = cv2.GaussianBlur(frame, (5, 5), 1)

    # trimming and grayscale conversion
    gray = cv2.cvtColor(fil[ymin:ymax, xmin:xmax], cv2.COLOR_BGR2GRAY)
    
    # binarization processing (Otsu's binarization)
    # return value : threshold value and binarized image
    th_val, bin = cv2.threshold(gray, 0, 255, cv2.THRESH_OTSU)
    
    # edge detection processing (canny method)
    edges = cv2.Canny(gray, 50, 100)   # C1-205 : 220, 330
    
    # outline extraction processing
    # return value : contours(pixel information) and hierarchy(Hierarchical structure information)
    contours, hierarchy = cv2.findContours(edges, cv2.RETR_TREE, cv2.CHAIN_APPROX_SIMPLE, offset=(xmin, ymin))

    return gray, bin, edges, contours

def sound():
    pygame.mixer.music.play()

fontType = cv2.FONT_HERSHEY_COMPLEX
obj_dis = 'xcm'

cap = cv2.VideoCapture(0) # mac:2  USB camera:0+cv2.CAP_DSHOW
cap.set(cv2.CAP_PROP_FOURCC, cv2.VideoWriter_fourcc('M', 'J', 'P', 'G'))
cap.set(cv2.CAP_PROP_FPS, 15) # FPS setting
cap.set(cv2.CAP_PROP_FRAME_WIDTH, 540) # width setting  1280pxel
cap.set(cv2.CAP_PROP_FRAME_HEIGHT, 360) # height setting  720pxel

fps = int(cap.get(cv2.CAP_PROP_FPS))
width = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
height = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
fourcc = cv2.VideoWriter_fourcc('m', 'p', '4', 'v')
video = cv2.VideoWriter('' + obj_dis + '.mp4', fourcc, fps, (width, height))

wb = openpyxl.load_workbook('xlsx')
wb.create_sheet(obj_dis)
ws = wb[obj_dis]

# trimming size setting
xmin, xmax = 240, 400
ymin, ymax = 180, 240

avg = None

white_list = []
time_list = []
val = 1
num = 4
base_time = time.time()
blink_time = 0

schedule.every(1.25).seconds.do(sound)

while True:
    schedule.run_pending()
    ret, frame = cap.read()
    if not ret:
        break

    gray, bin, edges, contours = filter(frame, xmin, xmax, ymin, ymax)
    
    if avg is None:
        avg = gray.copy().astype("float")
        continue

    cv2.accumulateWeighted(gray, avg, 0.6)
    frameDelta = cv2.absdiff(gray, cv2.convertScaleAbs(avg))

    thresh = cv2.threshold(frameDelta, 3, 255, cv2.THRESH_BINARY)[1]
    white_ratio, black_ratio = judge_square(thresh, xmin, xmax, ymin, ymax)
    contours, hierarchy = cv2.findContours(thresh.copy(), cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE, offset=(xmin, ymin))
    if white_ratio < 24 and white_ratio > 20:
        time_diff = time.time() - blink_time
        if time_diff > 0.3:
            blink_time = time.time()
            detection_time = blink_time - base_time
            cv2.putText(frame, 'Blink!', (260, 150), fontType, 1, (0, 0, 255), 3)
            ws.cell(row=num, column=4, value=val)
            ws.cell(row=num, column=5, value=detection_time)
            ws.cell(row=num, column=6, value=white_ratio)

            val += 1
            num += 1
    
    cv2.putText(frame, str(val - 1), (440, 440), fontType, 1, (0, 0, 255), 3)
    cv2.rectangle(frame, (xmin, ymin), (xmax, ymax), (255, 0, 0), 2)

    cv2.imshow("Frame", frame)
    cv2.imshow('edges', edges)

    print('white[%]', white_ratio)
    print('black[%]', black_ratio)
    
    end_time = time.time()
    run_time = end_time - base_time
    
    white_list.append(white_ratio)
    time_list.append(run_time)

    video.write(frame)
    if run_time > 30:
        ws.cell(row=2, column=2, value=val)
        break

    if cv2.waitKey(1) & 0xFF == ord('q'):
        break

wb.save('.xlsx')
wb.close()
pygame.quit()
plt.plot(time_list, white_list)
plt.show()
video.release()
cap.release()
cv2.destroyAllWindows()