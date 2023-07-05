import openpyxl

def sheet_setting(sheetname, path, blink_list, detectime_list, detecwhite_list, time_list, white_list):
    """
    概要:測定データを作成したEXCELファイルに書き込む関数
    
    注意事項
        ・while文をbreakした後に使用する
            →break前にすると処理が重くなるため
        ・引数に使用する変数はwhile文開始前に定義
            →'sheetnama'や'path'は固有値で定義
              それ以外は空配列で定義後、処理中に値を挿入していく
    引数
        sheetname:新規ワークシート作成用
        path:データ書き込み用EXCELファイルのパス
        blink_list:検出した瞬き回数の配列
        detectime_list:瞬きを検出した時間の配列
        detecwhite_list:瞬きを検出した白面積比率の配列
        time_list:プログラム走査時間配列(グラフ描画用)
        white_list:プログラム走査時の白面積比率配列(グラフ描画用)

    戻り値
        workbook:書き込み済みのEXCELファイル(.xlsx)

    """

    workbook = openpyxl.load_workbook(path)
    workbook.create_sheet(sheetname)
    worksheet = workbook[sheetname]
    worksheet["B2"].value = sheetname
    worksheet["D3"].value = "Detections"
    worksheet["E3"].value = "time[s]"
    worksheet["F3"].value = "white ratio[%]"
    worksheet["G3"].value = "run time[s]"
    worksheet["H3"].value = "white ratio[%]"

    for i in range(0, len(blink_list)):
        worksheet.cell(i + 4, 4, value = blink_list[i])
        worksheet.cell(i + 4, 5, value = detectime_list[i])
        worksheet.cell(i + 4, 6, value = detecwhite_list[i])

    for i1 in range(0, time_list):
        worksheet.cell(i1 + 4, 7, value = time_list[i1])
        worksheet.cell(i1 + 4, 8, value = white_list[i1])

    return workbook

