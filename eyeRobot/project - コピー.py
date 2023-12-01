import cv2
from cv2 import threshold
import numpy as np
import openpyxl

# Function of drawing circle
def draw(video, point, radius):
    pts = (int(point[0]), int(point[1]))
    r = int(radius)
    cv2.circle(video, pts, r, (100, 255, 0), 2)
    cv2.circle(video, pts, 2, (0,0,255), 3)     #中心点を描画

# Function of getting center point that be detected circle
def getpoint(xcenter, ycenter):
    wm, hm = 540, 360
    x = int((xcenter * wm) / 1000 * width)
    y = int((ycenter * hm) / 1000 * height)

    return x, y

# Function of judging iris (black and white area ratio)
def judge(frame, x, y, width, height):
    imagesize = width * height
    trim_frame = cv2.cvtColor(frame[y:y+height, x:x+width], cv2.COLOR_RGB2GRAY)
    ret_a, trim = cv2.threshold(trim_frame, 0, 255, cv2.THRESH_OTSU)
    white = cv2.countNonZero(trim)
    black = imagesize - white
    white_ratio = (white/imagesize) * 100
    black_ratio = (black/imagesize) * 100

    return white_ratio, black_ratio

def excel_data(x, y, val):
    wb = openpyxl.load_workbook('/Users/nakanokota/Desktop/Book1.xlsx')
    ws = wb.active
    n = val + 2
    ws.cell(row=n, column=2, value=x)
    ws.cell(row=n, column=3, value=y)
    wb.save('/Users/nakanokota/Desktop/Book1.xlsx')
    wb.close()


# camera settings
cap = cv2.VideoCapture(0+cv2.CAP_DSHOW) # mac:2  USB camera:0+cv2.CAP_DSHOW
cap.set(cv2.CAP_PROP_FOURCC, cv2.VideoWriter_fourcc('M', 'J', 'P', 'G'))
cap.set(cv2.CAP_PROP_FPS, 15) # FPS setting
width = cap.set(cv2.CAP_PROP_FRAME_WIDTH, 540) # width setting  1280pxel
height = cap.set(cv2.CAP_PROP_FRAME_HEIGHT, 360) # height setting  720pxel

# trimming size setting
xmin, xmax = 240, 400
ymin, ymax = 180, 240

val = 1

while True:
    # get numpy array
    # return value : bool value and image array
    bool, frame = cap.read()

    # GaussianFilter processing (smoothing)  ← variable
    filter = cv2.GaussianBlur(frame, (5, 5), 1)

    # trimming and grayscale conversion
    gray = cv2.cvtColor(filter[ymin:ymax, xmin:xmax], cv2.COLOR_RGB2GRAY)

    # binarization processing (Otsu's binarization)
    # return value : threshold value and binarized image
    threshold_value, bin = cv2.threshold(gray, 65, 255, cv2.THRESH_BINARY)

    # edge detection processing (canny method)  ← variable
    edges = cv2.Canny(gray, 100, 160)   # C1-205 : 220, 330

    # outline extraction processing
    # return value : contours(pixel information) and hierarchy(Hierarchical structure information)
    contours, hierarchy = cv2.findContours(edges, cv2.RETR_TREE, cv2.CHAIN_APPROX_SIMPLE, offset=(xmin, ymin))

    # enumerate contours (get index number and element)
    for i, cnt in enumerate(contours) :

        # minimum circumscribed circle detection process
        # return value : center(center point of the detected circle) and radius(radius of the detected circle) [pixel]
        center, radius = cv2.minEnclosingCircle(cnt)

        # minimum bounding rectangle detection process
        # return value : x, y, width, height (point[pixel])
        x, y, w, h = cv2.boundingRect(cnt)

        # iris determination processing
        # return value : white area ratio and black area ratio [%]
        white_ratio, black_ratio = judge(frame, x, y, w, h)

        # white area ratio < 30 , black area ratio < 70 [%]  ← variable
        if white_ratio < 30 and black_ratio > 70:

            # 10 < radius < 17 [pixel]  ← variable
            if radius < 17 and radius > 10:
                # circle drawing process
                draw(frame, (center[0], center[1]), radius)

                # numpy array conversion
                intarray = np.asarray(center, dtype = int)

                # point acquisition process
                x, y = getpoint(intarray[0], intarray[1])

                # input excel data for data sheet
                wb = openpyxl.load_workbook('.xlsx')  # excel file name
                ws = wb.active
                n = val + 3
                ws.cell(row=n, column=2, value=val)
                ws.cell(row=n, column=3, value=x)
                ws.cell(row=n, column=4, value=y)
                ws.cell(row=n, column=5, value=white_ratio)
                ws.cell(row=n, column=5, value=black_ratio)
                # ws.cell(row=n, column=5, value=threshold_value)
                wb.save('.xlsx')
                wb.close()
                val += 1
                
                # acquisition value confirmation
                # print('( x , y ) = ', x , y )
                # print('White Area [%] :', white_ratio)
                # print('Black Area [%] :', black_ratio)

        break

    # rectangle drawing process
    cv2.rectangle(frame, (xmin, ymin), (xmax, ymax), (255, 0, 0), 2)
    
    # displays
    cv2.imshow('output', frame)
    cv2.imshow('edge', edges)
    cv2.imshow('bin', bin)

    # Break by pressing "q" key
    if cv2.waitKey(1) & 0xFF == ord('q'):
        break

# freeing shooting objects and windows
cap.release()
cv2.destroyAllWindows()